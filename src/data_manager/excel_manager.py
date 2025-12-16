import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..config import Config

class SMCLabInfoManager:
    def __init__(self, config: Config = None):
        if config is None:
            config = Config()
        self.incre_data_path = config.incre_data_path
        self.filepath = config.info_plus_path
        try:
            self.df = pd.read_excel(self.filepath)
        except Exception as e:
            raise FileNotFoundError(f"请先运行SMCLabAddressbookParser")

    def map_fields(self, field1: str, field2: str):
        """
        建立两个字段之间的映射关系。
        字段包括: 
        - 姓名	
        - 年级	
        - 导师	
        - 培养类型(该项还没做好一致性检查)
        - 在读情况	
        - 飞书账号: ou_1df99022ddb02b52947cd7a76f42df3b
        - 学号	
        - union_id	
        - user_id
        - 邮箱	
        - 电话
        - 导师user_id	
        - 部门
        - 需要考勤

        返回：
            mapping_dict: {field1_value: field2_value}
            missing_names: list[姓名]
        """
        if field1 not in self.df.columns or field2 not in self.df.columns:
            raise ValueError(f"字段名错误：'{field1}' 或 '{field2}' 不存在于Excel中, 请检查'SMCLab学生扩展信息.xlsx'中是否有学生姓名不一致")

        mapping_dict = {} # 映射结果
        missing_names = [] # 统计缺失这两个字段的人员的姓名
        one2one_flag = True # 该映射是否是一对一映射

        for _, row in self.df.iterrows():
            val1, val2 = row[field1], row[field2]
            name = row.get("姓名", None)

            # 判断是否有缺失
            if pd.isna(val1) or pd.isna(val2):
                missing_names.append(name)
                continue  # 跳过这行，不加入映射

            # 若 val1 已存在，则追加到列表
            if val1 in mapping_dict:
                # 若不是列表则转为列表
                if not isinstance(mapping_dict[val1], list):
                    mapping_dict[val1] = [mapping_dict[val1]]
                    one2one_flag = False
                # 仅当 val2 不重复时才添加
                if val2 not in mapping_dict[val1]:
                    mapping_dict[val1].append(val2)
            else:
                mapping_dict[val1] = val2

        return mapping_dict, missing_names, one2one_flag

    def export_signature_sheet(self, output_path: str = None):
        """
        根据"需要考勤"列生成签名表，格式适合一页打印
        
        Args:
            output_path (str): 输出文件路径，如果为None则使用默认路径
        """
        # 检查"需要考勤"列是否存在
        if "需要考勤" not in self.df.columns:
            raise ValueError("DataFrame中不存在'需要考勤'列")
        
        # 筛选需要考勤的人员（需要考勤为1）
        attendance_df = self.df[self.df["需要考勤"] == 1].copy()
        
        if attendance_df.empty:
            raise ValueError("没有需要考勤的人员（'需要考勤'为1的行为空）")
        
        # 按姓名排序
        attendance_df = attendance_df.sort_values(by="姓名")
        
        # 设置输出路径
        if output_path is None:
            output_path = os.path.join(self.incre_data_path, "SMCLab在校签名表.xlsx")
        
        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "签名表"
        
        # 将人员列表分成两半
        total_count = len(attendance_df)
        mid_point = (total_count + 1) // 2  # 向上取整，确保左栏可以多一个人
        left_half = attendance_df.iloc[:mid_point]
        right_half = attendance_df.iloc[mid_point:]
        
        # 设置表头（两栏：左栏和右栏）
        headers = ["序号", "姓名", "签名", "序号", "姓名", "签名"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 设置样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # 填充左栏数据
        for idx, (_, row) in enumerate(left_half.iterrows(), start=1):
            row_num = idx + 1
            # 序号
            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=1).alignment = center_alignment
            ws.cell(row=row_num, column=1).border = thin_border
            
            # 姓名
            name = row.get("姓名", "")
            ws.cell(row=row_num, column=2, value=name)
            ws.cell(row=row_num, column=2).alignment = left_alignment
            ws.cell(row=row_num, column=2).border = thin_border
            
            # 签名栏（留空）
            ws.cell(row=row_num, column=3, value="")
            ws.cell(row=row_num, column=3).alignment = center_alignment
            ws.cell(row=row_num, column=3).border = thin_border
        
        # 填充右栏数据
        for idx, (_, row) in enumerate(right_half.iterrows(), start=1):
            row_num = idx + 1
            right_idx = mid_point + idx  # 右栏的序号从mid_point+1开始
            # 序号
            ws.cell(row=row_num, column=4, value=right_idx)
            ws.cell(row=row_num, column=4).alignment = center_alignment
            ws.cell(row=row_num, column=4).border = thin_border
            
            # 姓名
            name = row.get("姓名", "")
            ws.cell(row=row_num, column=5, value=name)
            ws.cell(row=row_num, column=5).alignment = left_alignment
            ws.cell(row=row_num, column=5).border = thin_border
            
            # 签名栏（留空）
            ws.cell(row=row_num, column=6, value="")
            ws.cell(row=row_num, column=6).alignment = center_alignment
            ws.cell(row=row_num, column=6).border = thin_border
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8  # 左栏序号
        ws.column_dimensions['B'].width = 18  # 左栏姓名
        ws.column_dimensions['C'].width = 25  # 左栏签名
        ws.column_dimensions['D'].width = 8  # 右栏序号
        ws.column_dimensions['E'].width = 18  # 右栏姓名
        ws.column_dimensions['F'].width = 25  # 右栏签名
        
        # 设置行高
        max_rows = max(len(left_half), len(right_half)) + 1
        ws.row_dimensions[1].height = 25  # 表头行高
        for row_idx in range(2, max_rows + 1):
            ws.row_dimensions[row_idx].height = 30  # 数据行高
        
        # 设置打印区域和页面设置（适合一页打印）
        ws.print_area = f'A1:F{max_rows}'
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = True
        ws.page_setup.fitToWidth = True
        
        # 保存文件
        wb.save(output_path)
        print(f"签名表已保存至: {output_path}")
        print(f"共 {len(attendance_df)} 人需要考勤")
    

if __name__ == "__main__":
    infomanager = SMCLabInfoManager()
    infomanager.export_signature_sheet()