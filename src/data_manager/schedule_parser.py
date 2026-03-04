import json, os, glob
import pandas as pd
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ..common.baseparser import SMCLabBaseParser
from ..config import Config

# 上午/下午/晚上 时段底色（浅绿→深绿）
PERIOD_FILLS = {
    "上午": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),  # 浅绿
    "下午": PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid"),  # 中绿
    "晚上": PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),  # 深绿
}
# 人数单元格：从浅蓝到深蓝的 RGB（用于按值插值）
COUNT_COLOR_LIGHT = (0xE7, 0xF5, 0xFE)  # #E7F5FE
COUNT_COLOR_DEEP = (0x87, 0xCE, 0xFA)   # #87CEFA

class SMCLabScheduleParser(SMCLabBaseParser):
    def __init__(self, config: Config = None):
        if config is None:
            config = Config()
        super().__init__(config)
        self.raw_data_path = config.schedule.raw_path
        self.time_table, self.days = self._build_time_table()
        self.sem_path = os.path.join(self._sem_data_path, self._year_semester)
        if not os.path.exists(self.sem_path):
            os.makedirs(self.sem_path, exist_ok=True)
    
    def _build_time_table(self):
        time_table = {
            "上午": {"第1节": ("08:00", "08:55"), "第2节": ("08:55", "09:40"), 
                   "第3节": ("10:10", "11:05"), "第4节": ("11:05", "11:50")},
            "下午": {"第1节": ("14:20", "15:15"), "第2节": ("15:15", "16:00"), 
                   "第3节": ("16:30", "17:15"), "第4节": ("17:25", "18:10")},
            "晚上": {"第1节": ("19:00", "19:55"), "第2节": ("19:55", "20:50"), 
                   "第3节": ("20:50", "21:35")}
        }
        days = ["周一", "周二", "周三", "周四", "周五"]
        return time_table, days
    
    def _all_slots(self):
        """返回所有时段元组列表 [('上午','第1节'), ('上午','第2节'), ...]"""
        return [(p, s) for p in self.time_table for s in self.time_table[p]]

    def _collect_schedule(self):
        """提取每个工作日每节课的上课同学名单"""
        file_list = sorted(glob.glob(os.path.join(self.raw_data_path, "*schedule_raw*.json")))
        if not file_list:
            raise FileNotFoundError(f"未在 {self.raw_data_path} 中找到任何 schedule*.json 文件")
        raw_file = file_list[0]
        if not os.path.exists(raw_file):
            raise RuntimeError(f"未找到 {raw_file}, 请先运行 SMCLabScheduleCrawler 下载数据")
        with open(raw_file, 'r', encoding='utf-8') as f:
            data = json.load(f)["items"]

        schedule = {day: defaultdict(list) for day in self.days}
        for item in data:
            fields = item.get("fields", {})
            name = fields.get("姓名", [{}])[0].get("text", "未知")
            for day in self.days:
                slots = fields.get(day, [])
                for slot in slots:
                    if "（当天无课程" in slot or "(当天无课程" in slot:
                        continue
                    schedule[day][slot].append(name)
        return schedule

    def _count_fill(self, value: int, value_max: int) -> PatternFill:
        """按人数在浅蓝到深蓝之间插值得到底色。value_max 为 0 时用浅蓝。"""
        if value_max <= 0:
            ratio = 0.0
        else:
            ratio = value / value_max
        r = int(COUNT_COLOR_LIGHT[0] + (COUNT_COLOR_DEEP[0] - COUNT_COLOR_LIGHT[0]) * ratio)
        g = int(COUNT_COLOR_LIGHT[1] + (COUNT_COLOR_DEEP[1] - COUNT_COLOR_LIGHT[1]) * ratio)
        b = int(COUNT_COLOR_LIGHT[2] + (COUNT_COLOR_DEEP[2] - COUNT_COLOR_LIGHT[2]) * ratio)
        hex_color = f"{r:02X}{g:02X}{b:02X}"
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def make_schedule_count_xlsx(self):
        output_path = os.path.join(self.sem_path, "schedule_count.xlsx")
        schedule = self._collect_schedule()
        records = []
        row_periods = []  # 每行对应的时段（上午/下午/晚上）
        for period, sub in self.time_table.items():
            for section, (start, end) in sub.items():
                row = {"上课节段": f"{period}-{section}", "时间": f"{start}-{end}"}
                for day in self.days:
                    names = schedule[day].get(f"{period}-{section}", [])
                    row[day] = len(names)
                records.append(row)
                row_periods.append(period)
        df = pd.DataFrame(records)
        df.to_excel(output_path, index=False, engine="openpyxl")

        # 按值计算人数列的最大值，用于插值
        count_cols = self.days  # 周一～周五
        value_max = max((r[day] for r in records for day in count_cols), default=0)

        wb = load_workbook(output_path)
        ws = wb.active
        for i, period in enumerate(row_periods):
            excel_row = i + 2  # 表头占第 1 行
            period_fill = PERIOD_FILLS[period]
            ws.cell(row=excel_row, column=1).fill = period_fill  # 上课节段
            ws.cell(row=excel_row, column=2).fill = period_fill  # 时间
            for j, day in enumerate(count_cols):
                cell = ws.cell(row=excel_row, column=3 + j)
                cell.fill = self._count_fill(records[i][day], value_max)
        wb.save(output_path)
        self.logger.info("课时人数表已保存: %s", Path(output_path).absolute())

    def make_schedule_names_xlsx(self):
        output_path = os.path.join(self.sem_path, "schedule_names.xlsx")
        schedule = self._collect_schedule()
        records = []
        for period, sub in self.time_table.items():
            for section, (start, end) in sub.items():
                row = {"上课节段": f"{period}-{section}", "时间": f"{start}-{end}"}
                for day in self.days:
                    names = schedule[day].get(f"{period}-{section}", [])
                    row[day] = ",".join(names)
                records.append(row)
        df = pd.DataFrame(records)
        df.to_excel(output_path, index=False)
        self.logger.info("课时姓名表已保存: %s", Path(output_path).absolute())

    def make_schedule_by_slot_json(self):
        output_path = os.path.join(self.sem_path, "schedule_by_slot.json")
        schedule = self._collect_schedule()
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(schedule, f, ensure_ascii=False, indent=4)
        self.logger.info("每节课学生名单JSON已保存: %s", Path(output_path).absolute())

    def make_period_summary_json(self):
        output_path = os.path.join(self.sem_path, "schedule_by_period.json")
        schedule = self._collect_schedule()
        summary = {day: {"上午": set(), "下午": set(), "晚上": set()} for day in self.days}
        for day in self.days:
            for slot, names in schedule[day].items():
                period, _ = slot.split("-")
                summary[day][period].update(names)
        # 转换为列表形式
        summary = {d: {p: sorted(list(v)) for p, v in ps.items()} for d, ps in summary.items()}
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=4)
        self.logger.info("每日三时段学生名单JSON已保存: %s", Path(output_path).absolute())
