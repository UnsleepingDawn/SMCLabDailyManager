import json, os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from ..common.baseparser import SMCLabBaseParser
from ..config import Config

class SMCLabAddressBookParser(SMCLabBaseParser):
    def __init__(self, config: Config = None):
        if config is None:
            config = Config()
        super().__init__(config)
        # 已经有的, 没有也能立刻创建的
        self.raw_data_path = os.path.join(config.raw_data_path, "address_book_raw_data")
        # 需要提前处理的
        self.excel_path = os.path.join(config.incre_data_path, "SMCLab学生基本信息.xlsx")
        if not os.path.exists(self.raw_data_path):
            os.makedirs(self.raw_data_path, exist_ok=True)
        # 需要提前下载的
        self.address_book_path = os.path.join(self.raw_data_path, "address_book.json")
        self.group_info_path = config.da_group_info_path
        # 输出路径
        self.output_path = os.path.join(config.incre_data_path, "SMCLab学生扩展信息.xlsx")
        self.merged_df = None
    
    def _extract_users_from_bitable(self):
        members = []
        with open(self.address_book_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)
        for department, primary_members in json_data.items():
            for user in primary_members.get("primary_members", []):
                members.append({
                    "姓名": user.get("name", ""),
                    "union_id": user.get("union_id", ""),
                    "飞书账号": user.get("open_id", ""),
                    "user_id": user.get("user_id", ""),
                    "邮箱": user.get("email", ""),
                    "电话": user.get("mobile", ""),
                    "培养类型": user.get("cultivation", ""),
                    "导师user_id": user.get("mentor_id", ""),
                    "部门": department,
                    "需要考勤": 0
                })
        return pd.DataFrame(members)

    def _fetch_attendance_id(self):
        """从 group_info.json 读取并返回 group_users_name_list 字段"""
        assert os.path.exists(self.group_info_path), f"group_info.json 文件不存在, 请先调用SMCLabAttendanceCrawler的get_group_info"
        with open(self.group_info_path, "r", encoding="utf-8") as f:
            group_info = json.load(f)
        group_users_id_list = group_info.get("group_users_id_list", [])
        self.logger.info("读取考勤组信息成功, 共 %d 个成员", len(group_users_id_list))
        return group_users_id_list

    def merge_dataframe(self):
        """合并并标记冲突"""
        # 读多维表格的信息
        json_df = self._extract_users_from_bitable()
        self.logger.info("读取SMCLab扩展信息成功, 共 %d 行", len(json_df))
        # 读通讯录的信息
        assert os.path.exists(self.excel_path), "请首先运行SMCLabMemberInfoParser, parse_all方法"
        excel_df = pd.read_excel(self.excel_path)
        self.logger.info("读取SMCLab基本信息成功, 共 %d 行", len(excel_df))
        # 读需要打卡的人的信息
        group_users_id_list = self._fetch_attendance_id()

        merged_df = pd.merge(
            excel_df,
            json_df,
            on="飞书账号",
            how="outer",
            suffixes=("_Excel", "_JSON"),
            indicator=True
        )

        # 根据user_id将需要打卡的人的"需要考勤"字段设为1
        # 将group_users_id_list中的user_id对应的"需要考勤"设为1
        merged_df.loc[merged_df["user_id"].isin(group_users_id_list), "需要考勤"] = 1
        self.logger.info("已将 %d 个需要打卡的人员的'需要考勤'字段设为1", 
                        merged_df[merged_df["user_id"].isin(group_users_id_list)].shape[0])

        # 用 Excel 和 JSON 数据进行缺失值的相互填充以尽量补足姓名和培养类型
        for idx, row in merged_df.iterrows():
            if not pd.notna(row["姓名_JSON"]) and pd.notna(row["姓名_Excel"]):
                merged_df.at[idx, "姓名_JSON"] = row["姓名_Excel"]
            if not pd.notna(row["姓名_Excel"]) and pd.notna(row["姓名_JSON"]):
                merged_df.at[idx, "姓名_Excel"] = row["姓名_JSON"]
            if not pd.notna(row["培养类型_JSON"]) and pd.notna(row["培养类型_Excel"]):
                merged_df.at[idx, "培养类型_JSON"] = row["培养类型_Excel"]
            if not pd.notna(row["培养类型_Excel"]) and pd.notna(row["培养类型_JSON"]):
                merged_df.at[idx, "培养类型_Excel"] = row["培养类型_JSON"]


        conflict_cols = []
        # 对于姓名, 培养类型, 有冲突就记录, 没冲突就合并
        for col in ["姓名", "培养类型"]:
            col_excel = f"{col}_Excel"
            col_json = f"{col}_JSON"
            if col_excel in merged_df.columns and col_json in merged_df.columns:

                merged_df[f"{col}_冲突"] = merged_df[col_excel] != merged_df[col_json] 

                if merged_df[f"{col}_冲突"].any():
                    self.logger.warning("********* %s 字段有冲突！请优先处理！ *********", col)
                    conflict_cols.append(col)
                else:
                    merged_df.rename(columns={
                                            f"{col_excel}": f"{col}",
                                        }, inplace=True)
                    merged_df.drop(columns=col_json, inplace=True)
                    merged_df.drop(columns=f"{col}_冲突", inplace=True)

        self.merged_df = merged_df
        self.conflict_cols = conflict_cols
        self.merged_df.to_excel(self.output_path, index=False)
        self.logger.info("合并完成，共 %d 行，其中 %d 个字段存在冲突", len(merged_df), len(conflict_cols))

    def mark_info_in_excel(self, update: bool=True):
        """保存Excel并对冲突单元格标红加粗"""
        if not os.path.exists(self.output_path) or update:
            self.merge_dataframe()

        wb = load_workbook(self.output_path)
        ws = wb.active

        # 样式：红底、白字、加粗
        conflict_font = Font(bold=True, color="B01E1C")

        # 对冲突单元格进行标红加粗
        for col in self.conflict_cols:
            excel_col = f"{col}_Excel"
            json_col = f"{col}_JSON"
            conflict_flag = f"{col}_冲突"

            if conflict_flag not in self.merged_df.columns:
                continue

            for idx, is_conflict in enumerate(self.merged_df[conflict_flag], start=2):  # 从第二行开始（第一行为表头）
                if is_conflict:
                    excel_col_idx = self.merged_df.columns.get_loc(excel_col) + 1
                    json_col_idx = self.merged_df.columns.get_loc(json_col) + 1
                    ws.cell(idx, excel_col_idx).font = conflict_font
                    ws.cell(idx, json_col_idx).font = conflict_font
                    if '姓名' not in self.conflict_cols:
                        name_col_idx = self.merged_df.columns.get_loc('姓名') + 1
                    else:
                        name_col_idx = self.merged_df.columns.get_loc('姓名_Excel') + 1

                    name = ws.cell(idx, name_col_idx).value
                    self.logger.warning("\t%s\t的*%s*字段有冲突: %s vs. %s", name, col, ws.cell(idx, excel_col_idx).value, ws.cell(idx, json_col_idx).value)


        wb.save(self.output_path)
        self.logger.info("文件已保存：%s", self.output_path)
        self.logger.info("红色加粗部分表示Excel与JSON存在差异")


