import json, os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ABS_PATH = os.path.abspath(__file__)        # SMCLabDailyManager\source_code\SMCLabDataManager\AddressBookParser.py
CURRENT_PATH = os.path.dirname(ABS_PATH)    # SMCLabDailyManager\source_code\SMCLabDataManager
SRC_PATH = os.path.dirname(CURRENT_PATH)    # SMCLabDailyManager\source_code
REPO_PATH = os.path.dirname(SRC_PATH)       # SMCLabDailyManager
RAW_DATA_PATH = os.path.join(REPO_PATH, "data_raw") # SMCLabDailyManager\data_raw
INCRE_DATA_PATH = os.path.join(REPO_PATH, "data_incremental") # SMCLabDailyManager\data_incremental

class SMCLabAddressBookParser:
    def __init__(self):
        self.excel_path = os.path.join(INCRE_DATA_PATH, "SMCLab学生基本信息.xlsx")
        self.raw_data_path = os.path.join(RAW_DATA_PATH, "address_book_raw_data")
        if not os.path.exists(self.raw_data_path):
            os.makedirs(self.raw_data_path, exist_ok=True)
        self.address_book_path = os.path.join(self.raw_data_path, "address_book.json")
        self.output_path = os.path.join(INCRE_DATA_PATH, "SMCLab学生扩展信息.xlsx")
        self.df = self._read_excel()
        self.json_data = self._read_json()
        self.merged_df = None

    def _read_excel(self):
        # 用于init
        assert os.path.exists(self.excel_path), "请首先运行SMCLabMemberInfoParser, parse_all方法"

        df = pd.read_excel(self.excel_path)
        print(f"读取Excel成功, 共 {len(df)} 行")
        return df

    def _read_json(self):
        # 用于init
        with open(self.address_book_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)
        print(f"读取JSON成功")
        return json_data
    
    # def _index_by_open_id(self):
    #     """返回以飞书账号为索引的 DataFrame 副本（保证唯一索引）"""
    #     df = self.df
    #     if "飞书账号" not in df.columns:
    #         raise ValueError("需要包含列 '飞书账号' 才能索引")
    #     # 若存在重复 open_id，保留第一个并警告
    #     dupes = df["飞书账号"][df["飞书账号"] != ""].duplicated()
    #     if dupes.any():
    #         print("警告：存在重复的飞书账号(open_id)，将在合并时以第一个为准。")
    #     df_idx = df.set_index("飞书账号", drop=False)
    #     # 若有重复索引， keep first
    #     df_idx = df_idx[~df_idx.index.duplicated(keep="first")]
    #     self.df = df_idx
    
    def _extract_json_members(self):
        members = []
        for department, primary_members in self.json_data.items():
            for user in primary_members.get("primary_members", []):
                members.append({
                    "姓名": user.get("name", ""),
                    "union_id": user.get("union_id", ""),
                    "飞书账号": user.get("open_id", ""),
                    "user_id": user.get("user_id", ""),
                    "邮箱": user.get("email", ""),
                    "电话": user.get("mobile", ""),
                    "培养类型": user.get("cultivation", ""),
                    "导师user_id": user.get("mentor_id", ""),
                    "部门": department
                })
        return pd.DataFrame(members)
    
    # def _build_mentor_mapping(self, json_df):
    #     """根据JSON中的所有成员构建 user_id -> 姓名 映射"""
    #     mapping = {}
    #     for _, row in json_df.iterrows():
    #         if row["部门"] != "Tenure": continue
    #         uid = row.get("user_id")
    #         name = row.get("姓名")
    #         if pd.notna(uid) and uid:
    #             mapping[uid] = name
    #     return mapping
    
    def merge(self):
        """合并并标记冲突"""
        json_df = self._extract_json_members()
        excel_df = self.df.copy()

        merged = pd.merge(
            excel_df,
            json_df,
            on="飞书账号",
            how="outer",
            suffixes=("_Excel", "_JSON"),
            indicator=True
        )

        # 更新已有行
        for idx, row in merged.iterrows():
            if not pd.notna(row["姓名_JSON"]) and pd.notna(row["姓名_Excel"]):
                merged.at[idx, "姓名_JSON"] = row["姓名_Excel"]
            if not pd.notna(row["姓名_Excel"]) and pd.notna(row["姓名_JSON"]):
                merged.at[idx, "姓名_Excel"] = row["姓名_JSON"]
            if not pd.notna(row["培养类型_JSON"]) and pd.notna(row["培养类型_Excel"]):
                merged.at[idx, "培养类型_JSON"] = row["培养类型_Excel"]
            if not pd.notna(row["培养类型_Excel"]) and pd.notna(row["培养类型_JSON"]):
                merged.at[idx, "培养类型_Excel"] = row["培养类型_JSON"]

        conflict_cols = []
        # 对于姓名, 培养类型, 有冲突就记录, 没冲突就合并
        for col in ["姓名", "培养类型"]:
            col_excel = f"{col}_Excel"
            col_json = f"{col}_JSON"
            if col_excel in merged.columns and col_json in merged.columns:

                merged[f"{col}_冲突"] = merged[col_excel] != merged[col_json] 

                if merged[f"{col}_冲突"].any():
                    print(f"********* {col} 字段有冲突！请优先处理！ *********")
                    conflict_cols.append(col)
                else:
                    merged.rename(columns={
                                            f"{col_excel}": f"{col}",
                                        }, inplace=True)
                    merged.drop(columns=col_json, inplace=True)
                    merged.drop(columns=f"{col}_冲突", inplace=True)

        self.merged_df = merged
        self.conflict_cols = conflict_cols
        print(f"合并完成，共 {len(merged)} 行，其中 {len(conflict_cols)} 个字段存在冲突")

    def merge_info_to_excel(self):
        """保存Excel并对冲突单元格标红加粗"""
        if not self.merged_df:
            self.merge()
        output_path = self.output_path
        self.merged_df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        # 样式：红底、白字、加粗
        conflict_font = Font(bold=True, color="B01E1C")

        # 对冲突单元格进行标红加粗
        for col in self.conflict_cols:
            excel_col = f"{col}_Excel"
            json_col = f"{col}_JSON"
            conflict_flag = f"{col}_冲突"

            if conflict_flag not in self.merged_df.columns:
                continue

            for idx, is_conflict in enumerate(self.merged_df[conflict_flag], start=2):  # 从第二行开始（第一行为表头）
                if is_conflict:
                    excel_col_idx = self.merged_df.columns.get_loc(excel_col) + 1
                    json_col_idx = self.merged_df.columns.get_loc(json_col) + 1
                    ws.cell(idx, excel_col_idx).font = conflict_font
                    ws.cell(idx, json_col_idx).font = conflict_font


        wb.save(output_path)
        print(f"文件已保存：{output_path}")
        print("红色加粗部分表示Excel与JSON存在差异")

# ---------------- 使用示例 ----------------
if __name__ == "__main__":
    parser = SMCLabAddressBookParser()
    parser.merge_info_to_excel()
